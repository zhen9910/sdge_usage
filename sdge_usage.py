import openpyxl
import holidays
import datetime
import unittest
import sys
import math

us_holidays = holidays.US()

def get_time_periods(date, time):

    hour = time.hour
    weekday = date.weekday() < 5
    weekend = not weekday
    holiday = date in us_holidays
    march_or_april = date.month == 3 or date.month == 4

    match hour:
        # case 0 | 1 | 2 | 3 | 4 | 5:
        case hour if hour >=0 and hour < 6:
            time_period = "Super Off Peak"
        # case 6 | 7 | 8 | 9:
        case hour if hour >= 6 and hour < 10:
            time_period = "Super Off Peak" if (weekend or holiday) else "Off Peak"
        # case 10 | 11 | 12 | 13:
        case hour if hour >= 10 and hour < 14:
            time_period = "Super Off Peak" if (weekend or holiday or march_or_april) else "Off Peak"
        # case 16 | 17 | 18 | 19, 20:
        case hour if hour >= 16 and hour < 21:
            time_period = "On Peak"
        case _:
            time_period = "Off Peak"

    # print(f"{time_str} is at {hour}")
    # day_type = "weekday" if weekday else "weekend"
    # holiday_type = "holiday" if holiday else "not a holiday"
    # march_or_april_type = "in" if march_or_april else "not March or April"
    # print(f"{date_str} is a {day_type} {holiday_type}")
    # print(f"{date_str} is {march_or_april_type} March or April")
    # print(f"{date_str} {time_str} is at {time_period}")
    # print("\n\n")

    return time_period

if __name__ == '__main__':

    if len(sys.argv) != 2:
        print("Usage: python sdge_usage.py filename")
        sys.exit(1)

    # filename = "Electric_15_Minute_3-14-2023_4-5-2023_20230406.xlsx"
    # filename = "Electric_15_Minute_2-10-2023_3-13-2023_20230406.xlsx"
    filename = sys.argv[1]

    # open file
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    super_off_peak_kwh = 0
    off_peak_kwh = 0
    on_peak_kwh = 0
    meter_number = 0

    # get metadata
    for row in sheet.iter_rows(values_only=True):
        match row[0]:
            case "Meter Number":
                if isinstance(row[1], int):
                    meter_number = row[1]
                else:
                    break
            case "Reading Start":
                reading_start = row[1]
            case "Reading End":
                reading_end = row[1]
            case "Total Usage":
                total_usage = row[1]

    # get usage data
    for row in sheet.iter_rows(values_only=True):
            if row[0] == meter_number:
                # print(row)
                date = row[1]
                time = row[2]
                kwh = row[6]
                time_period = get_time_periods(date, time)
                match time_period:
                    case "Super Off Peak":
                        super_off_peak_kwh += kwh
                    case "Off Peak":
                        off_peak_kwh += kwh
                    case "On Peak":
                        on_peak_kwh += kwh

    total_kwh = super_off_peak_kwh + off_peak_kwh + on_peak_kwh

    if not math.isclose(total_kwh, total_usage):
        print(f"total usage is incorrect: {total_kwh} != {total_usage}")
        sys.exit(1)

    print(f"\n\nElectricity usage from {reading_start} to {reading_end}\n\n")
    print(f"Super Off Peak net usage (kwh): {super_off_peak_kwh}")
    print(f"Off Peak net usage (kwh):       {off_peak_kwh}")
    print(f"On Peak net usage (kwh):        {on_peak_kwh}\n")

    print(f"Total net usage (kwh):          {total_kwh}")
    # print(f"total usage is:        {total_usage}")
